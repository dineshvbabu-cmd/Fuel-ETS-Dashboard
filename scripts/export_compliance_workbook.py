from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(r"c:\Users\DineshBabu\Downloads\EU_ETS_FuelEU_Compliance_Calculator_10_1_1.xlsx")
SOURCE = Path(os.environ.get("COMPLIANCE_WORKBOOK_PATH", DEFAULT_SOURCE))
OUTPUT = ROOT / "compliance_dashboard" / "data" / "workbook-seed.json"


CALCULATOR_EDITABLE_HEADERS = [
    "IMO No.\n(enter → auto-fill)",
    "Departure Date",
    "From Port\nUN/LOCODE",
    "Arrival Date",
    "To Port\nUN/LOCODE",
    "Fossil Fuel 1\nType",
    "Fossil Fuel 1\nCons. (MT)",
    "Fossil Fuel 2\nType",
    "Fossil Fuel 2\nCons. (MT)",
    "Biofuel/RFNBO\nType",
    "Biofuel/RFNBO\nCons. (MT)",
    "Sustain. Factor\nWtW (0–1)",
    "WASP Factor\n(f_wind)",
    "Total Distance\n(nm)",
    "Cargo Carried\n(t)",
    "Time at Sea\n(h)",
    "Hours at\nBerth/Anchor",
    "OPS Electricity\n(MJ)",
]


PARAMETER_ROWS = [
    {"section": "EU ETS — Emissions Trading System", "key": "reportYear", "row": 5, "editable": True, "type": "number"},
    {"section": "EU ETS — Emissions Trading System", "key": "etsPhaseIn", "row": 6, "editable": False, "type": "number"},
    {"section": "EU ETS — Emissions Trading System", "key": "etsGasScope", "row": 7, "editable": False, "type": "text"},
    {"section": "EU ETS — Emissions Trading System", "key": "bioZero", "row": 8, "editable": True, "type": "text"},
    {"section": "EU ETS — Emissions Trading System", "key": "euaPrice", "row": 9, "editable": True, "type": "number"},
    {"section": "FuelEU Maritime", "key": "fueleuRef", "row": 12, "editable": True, "type": "number"},
    {"section": "FuelEU Maritime", "key": "fueleuRedux", "row": 13, "editable": False, "type": "number"},
    {"section": "FuelEU Maritime", "key": "fueleuTarget", "row": 14, "editable": False, "type": "number"},
    {"section": "FuelEU Maritime", "key": "vlsfoMj", "row": 15, "editable": True, "type": "number"},
    {"section": "FuelEU Maritime", "key": "penRate", "row": 16, "editable": True, "type": "number"},
    {"section": "FuelEU Maritime", "key": "rfnboWindow", "row": 17, "editable": True, "type": "text"},
    {"section": "Global Warming Potentials", "key": "gwpCo2", "row": 20, "editable": True, "type": "number"},
    {"section": "Global Warming Potentials", "key": "gwpCh4", "row": 21, "editable": True, "type": "number"},
    {"section": "Global Warming Potentials", "key": "gwpN2o", "row": 22, "editable": True, "type": "number"},
    {"section": "Global Warming Potentials", "key": "gwpBasis", "row": 23, "editable": True, "type": "text"},
    {"section": "Global Warming Potentials", "key": "penMultiplier", "row": 24, "editable": True, "type": "number"},
    {"section": "Global Warming Potentials", "key": "elecWtw", "row": 25, "editable": True, "type": "number"},
    {"section": "Global Warming Potentials", "key": "gwpCh4Ets", "row": 26, "editable": True, "type": "number"},
    {"section": "Global Warming Potentials", "key": "gwpN2oEts", "row": 27, "editable": True, "type": "number"},
]


def clean_header(value, fallback):
    if value is None or value == "":
        return fallback
    return str(value)


def read_table(ws, header_row: int, start_row: int, end_row: int | None = None):
    headers = [clean_header(ws.cell(header_row, column).value, f"Column {column}") for column in range(1, ws.max_column + 1)]
    if end_row is None:
        end_row = ws.max_row
    rows = []
    for row_index in range(start_row, end_row + 1):
        values = [ws.cell(row_index, column).value for column in range(1, ws.max_column + 1)]
        if any(value is not None and value != "" for value in values):
            rows.append(values)
    return headers, rows


def rows_to_objects(headers, rows):
    objects = []
    for row in rows:
        record = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            record[str(header)] = row[index] if index < len(row) else None
        objects.append(record)
    return objects


def export():
    wb_values = load_workbook(SOURCE, data_only=True)

    calculator_ws = wb_values["Calculator"]
    calculator_headers, calculator_rows = read_table(calculator_ws, 4, 5)
    calculator_objects = rows_to_objects(calculator_headers, calculator_rows)

    calculator_seed = []
    for row in calculator_objects:
        calculator_seed.append(
            {
                "imoNo": row.get("IMO No.\n(enter → auto-fill)"),
                "departureDate": row.get("Departure Date"),
                "fromPortCode": row.get("From Port\nUN/LOCODE"),
                "arrivalDate": row.get("Arrival Date"),
                "toPortCode": row.get("To Port\nUN/LOCODE"),
                "fuel1Type": row.get("Fossil Fuel 1\nType"),
                "fuel1ConsumptionMt": row.get("Fossil Fuel 1\nCons. (MT)"),
                "fuel2Type": row.get("Fossil Fuel 2\nType"),
                "fuel2ConsumptionMt": row.get("Fossil Fuel 2\nCons. (MT)"),
                "bioFuelType": row.get("Biofuel/RFNBO\nType"),
                "bioFuelConsumptionMt": row.get("Biofuel/RFNBO\nCons. (MT)"),
                "sustainabilityFactor": row.get("Sustain. Factor\nWtW (0–1)"),
                "windFactor": row.get("WASP Factor\n(f_wind)"),
                "distanceNm": row.get("Total Distance\n(nm)"),
                "cargoTonnes": row.get("Cargo Carried\n(t)"),
                "timeAtSeaHours": row.get("Time at Sea\n(h)"),
                "berthHours": row.get("Hours at\nBerth/Anchor"),
                "opsElectricityMj": row.get("OPS Electricity\n(MJ)"),
            }
        )

    parameters_ws = wb_values["Parameters"]
    parameter_rows = []
    for definition in PARAMETER_ROWS:
        parameter_rows.append(
            {
                "section": definition["section"],
                "key": definition["key"],
                "label": parameters_ws.cell(definition["row"], 1).value,
                "value": parameters_ws.cell(definition["row"], 2).value,
                "note": parameters_ws.cell(definition["row"], 3).value,
                "editable": definition["editable"],
                "type": definition["type"],
            }
        )

    fuel_headers, fuel_rows = read_table(wb_values["Fuel_Reference"], 4, 5)
    fuel_reference = rows_to_objects(fuel_headers, fuel_rows)

    fleet_headers, fleet_rows = read_table(wb_values["Fleet_DB"], 4, 5)
    fleet = rows_to_objects(fleet_headers, fleet_rows)

    port_headers, port_rows = read_table(wb_values["Port_DB"], 4, 5)
    ports = rows_to_objects(port_headers, port_rows)

    flag_headers, flag_rows = read_table(wb_values["Flag_States"], 4, 5)
    flags = rows_to_objects(flag_headers, flag_rows)

    derogation_headers, derogation_rows = read_table(wb_values["Derogations"], 7, 8)
    derogations = rows_to_objects(derogation_headers, derogation_rows)

    methodology_ws = wb_values["Methodology"]
    methodology = []
    for row_index in range(1, methodology_ws.max_row + 1):
        detail = methodology_ws.cell(row_index, 2).value
        if detail:
            methodology.append({"detail": detail})

    formula_headers, formula_rows = read_table(wb_values["Formula_Guide"], 2, 3)
    formula_guide = rows_to_objects(formula_headers, formula_rows)

    payload = {
        "generatedAt": datetime.now().isoformat(),
        "sourceWorkbook": str(SOURCE),
        "sheetOrder": [
            "Dashboard",
            "Calculator",
            "Vessel Summary",
            "Parameters",
            "Fuel Reference",
            "Fleet DB",
            "Port DB",
            "Flag States",
            "Derogations",
            "Methodology",
            "Formula Guide",
        ],
        "calculator": {
            "editableHeaders": list(CALCULATOR_EDITABLE_HEADERS),
            "rows": calculator_seed,
        },
        "parameters": parameter_rows,
        "fuelReference": fuel_reference,
        "fleet": fleet,
        "ports": ports,
        "flags": flags,
        "derogations": derogations,
        "methodology": methodology,
        "formulaGuide": formula_guide,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    export()
